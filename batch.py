import json
import os
from datetime import date
from agent import run_agent, client
import openpyxl

os.makedirs("outputs", exist_ok=True)
for file in os.listdir("outputs"):
    if file.endswith(".md"):
        os.remove(os.path.join("outputs", file))
        
def save_to_tracker(result, url, filepath="tracker.xlsx"):
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Company", "Position", "Industry", "Fit Score", "Decision", "Matches", "Gaps", "Reasoning", "CV: Add", "CV: Remove", "URL", "Date"])    
        
    ws.append([
        result.get("company", ""),
        result.get("position", ""),
        result.get("industry", ""),
        result.get("fit_score", ""),
        result.get("decision", ""),
        ", ".join(result.get("matches", [])),
        ", ".join(result.get("gaps", [])),
        result.get("reasoning", ""),
        ", ".join(result.get("cv_recommendations", {}).get("add", [])),
        ", ".join(result.get("cv_recommendations", {}).get("remove", [])),
        url,
        str(date.today())
    ])  
    wb.save(filepath)

def generate_cv_suggestions(result, cv_path="data/cv.md"):
    with open(cv_path, "r") as f:
        cv_content = f.read()
    
    response = client.chat.completions.create(
        model="gpt-4o",
        temperature=0,
        messages=[
            {
                "role": "system",
                "content": """You are an expert CV advisor.
                Given a CV and a job, suggest specific targeted changes only.
                Rules:
                - Only suggest changes where there is a clear reason
                - Never invent experience that isn't in the CV
                - For rewrites, show original bullet and suggested version
                - Be concise and specific
                - Return as clean markdown"""
            },
            {
                "role": "user",
                "content": f"""
                CV:
                {cv_content}
                
                Job: {result.get('position')} at {result.get('company')}
                Add to CV: {', '.join(result.get('cv_recommendations', {}).get('add', []))}
                Remove from CV: {', '.join(result.get('cv_recommendations', {}).get('remove', []))}
                Gaps to address: {', '.join(result.get('gaps', []))}
                
                Return markdown with these sections:
                ## Add to CV
                ## Remove from CV  
                ## Suggested rewrites
                """
            }
        ]
    )
    return response.choices[0].message.content

def save_cv_suggestions(content, company):
    company_slug = company.replace(" ", "_").lower()
    filename = f"outputs/{company_slug}_cv_suggestions.md"
    with open(filename, "w") as f:
        f.write(content)
    return filename

def already_tracked(url, filepath="tracker.xlsx"):
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if url in row:
                return True
    except FileNotFoundError:
        pass
    return False

with open("data/jobs/urls.txt", "r") as url_file:
    urls = [line.strip() for line in url_file if line.strip()]

for i, url in enumerate(urls):
    print(f"\nAnalyzing job {i+1}/{len(urls)}: {url}")
    if already_tracked(url):
        print(f"↷ Already in tracker, skipping")
        continue
    try:
        result = run_agent(url)
        
        company_name = result.get('company', f'job_{i+1}').replace(" ", "_").lower()
        final_decision = "apply" if result.get("fit_score", 0) >= 70 else "skip"

        with open(f"outputs/{company_name}.md", "w") as f:
            f.write(f"# {result.get('company')} — {result.get('position')}\n\n")
            f.write(f"**Fit Score:** {result.get('fit_score')}/100\n")
            f.write(f"**Decision:** {final_decision.upper()}\n\n")
            f.write(f"## Reasoning\n{result.get('reasoning', '')}\n\n")
            f.write(f"## Matches\n")
            for m in result.get("matches", []):
                f.write(f"- {m}\n")
            f.write(f"\n## Gaps\n")
            for g in result.get("gaps", []):
                f.write(f"- {g}\n")
            if final_decision == "apply":
                f.write(f"\n## CV Recommendations\n")
                f.write(f"**Add:** {', '.join(result.get('cv_recommendations', {}).get('add', []))}\n")
                f.write(f"**Remove:** {', '.join(result.get('cv_recommendations', {}).get('remove', []))}\n")

        if final_decision == "apply":
            save_to_tracker(result, url)
            cv_suggestions = generate_cv_suggestions(result)
            cv_file = save_cv_suggestions(cv_suggestions, result.get('company', ''))
            print(f"✓ {result.get('company')} — {result.get('fit_score')}/100 — APPLY → saved to tracker")
            print(f"  CV suggestions: {cv_file}")
        else:
            print(f"✗ {result.get('company')} — {result.get('fit_score')}/100 — SKIP")
        
    except Exception as e:
        import traceback
        traceback.print_exc()