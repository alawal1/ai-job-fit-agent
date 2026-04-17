from dashbord_part1 import normalize_text, split_tags, parse_markdown_report, OUTPUTS_DIR, TRACKER_PATH, DASHBOARD_PATH
from openpyxl import load_workbook
import os


def load_reports():
    reports = []
    if not os.path.isdir(OUTPUTS_DIR):
        return reports
    for filename in sorted(os.listdir(OUTPUTS_DIR)):
        if not filename.lower().endswith('.md'):
            continue
        if filename.lower().endswith('_cv.md'):
            continue
        if '.' in filename[:-3]:
            continue
        path = os.path.join(OUTPUTS_DIR, filename)
        reports.append(parse_markdown_report(path))
    return reports


def load_tracker(path):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}

    headers = [normalize_text(cell).lower() for cell in rows[0]]
    tracker = {}
    for row in rows[1:]:
        if not any(cell is not None for cell in row):
            continue
        row_data = {headers[i]: row[i] for i in range(len(headers)) if headers[i]}
        company = normalize_text(row_data.get('company'))
        position = normalize_text(row_data.get('position'))
        if 'cv improvements' in company.lower() or 'cv improvements' in position.lower():
            continue
        if not company or not position:
            continue

        score = row_data.get('fit score') if 'fit score' in row_data else row_data.get('fit_score')
        try:
            score = int(score)
        except Exception:
            score = None

        key = f"{company.lower()}||{position.lower()}"
        tracker[key] = {
            'company': company,
            'position': position,
            'fit_score': score,
            'decision': normalize_text(row_data.get('decision')),
            'matches': split_tags(row_data.get('matches')),
            'gaps': split_tags(row_data.get('gaps')),
        }
    return tracker


def match_jobs(reports, tracker):
    jobs = []
    for report in reports:
        key = f"{report['company'].lower()}||{report['position'].lower()}"
        tracked = tracker.get(key, {})
        jobs.append({
            'company': report['company'],
            'position': report['position'],
            'fit_score': report['fit_score'] if report['fit_score'] is not None else tracked.get('fit_score'),
            'decision': report['decision'] or tracked.get('decision', ''),
            'matches': report['matches'] or tracked.get('matches', []),
            'gaps': report['gaps'] or tracked.get('gaps', []),
            'snippet': report['snippet'],
        })
    return jobs


def score_color(score, decision):
    if isinstance(score, int):
        if score >= 70:
            return '#2e7d32'
        if 60 <= score < 70:
            return '#f9a825'
        return '#c62828'
    text = normalize_text(decision).lower()
    if 'apply' in text:
        return '#2e7d32'
    if 'skip' in text:
        return '#c62828'
    return '#616161'
