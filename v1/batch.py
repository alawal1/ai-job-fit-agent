# batch.py
import json
import os
from datetime import date
from agent import run_agent, client, fetch_job_posting
import openpyxl

os.makedirs("outputs", exist_ok=True)
for file in os.listdir("outputs"):
    if file.endswith(".md"):
        os.remove(os.path.join("outputs", file))
        
def save_to_tracker(result, url, filepath="tracker.xlsx"):
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Company", "Position", "Industry", "Fit Score", "Decision", "Matches", "Gaps", "Reasoning", "CV: Add", "CV: Remove", "URL", "Date"])    
        
    ws.append([
        result.get("company", ""),
        result.get("position", ""),
        result.get("industry", ""),
        result.get("fit_score", ""),
        result.get("decision", ""),
        ", ".join(result.get("matches", [])),
        ", ".join(result.get("gaps", [])),
        result.get("reasoning", ""),
        ", ".join(result.get("cv_recommendations", {}).get("add", [])),
        ", ".join(result.get("cv_recommendations", {}).get("remove", [])),
        url,
        str(date.today())
    ])  
    wb.save(filepath)


def improve_cv(result, cv_path, job_text):
    """Return side-by-side rewrite suggestions for each CV work experience bullet."""
    with open(cv_path, "r") as f:
        cv_content = f.read()

    response = client.chat.completions.create(
        model="gpt-4o",
        temperature=0,
        messages=[
            {
                "role": "system",
                "content": """You are an expert CV advisor focused on job alignment.
                Given a CV and a target job posting, suggest concise rewrite improvements for work experience bullets.
                Rules:
                - Prioritize inserting specific job keywords, skills, and requirements from the full job posting.
                - Keep bullets short and relevant rather than verbose.
                - Do not invent new accomplishments or change the original facts.
                - Show each bullet's original text and a suggested rewritten version side by side.
                - Return clean markdown with a clear structure.
                - never add verbs or actions that aren't in the original bullet
                """
            },
            {
                "role": "user",
                "content": f"""
                CV:
                {cv_content}

                Job: {result.get('position')} at {result.get('company')}
                Job posting:
                {job_text}

                Focus on the job's matches, gaps, and CV recommendations.
                Insert the most relevant keywords and responsibilities from the job posting into each bullet.

                Return markdown with a section titled "## Bullet rewrite suggestions".
                For each work experience bullet, include:
                - Original: <original bullet>
                - Suggested: <rewritten version>
                """
            }
        ]
    )
    return response.choices[0].message.content



def already_tracked(url, filepath="tracker.xlsx"):
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if url in row:
                return True
    except FileNotFoundError:
        pass
    return False

with open("data/jobs/urls.txt", "r") as url_file:
    urls = [line.strip() for line in url_file if line.strip()]

for i, url in enumerate(urls):
    print(f"\nAnalyzing job {i+1}/{len(urls)}: {url}")
    if already_tracked(url):
        print(f"↷ Already in tracker, skipping")
        continue
    try:
        job_text = fetch_job_posting(url)
        result = run_agent(url)
        
        company_name = result.get('company', f'job_{i+1}').replace(" ", "_").lower()
        final_decision = "apply" if result.get("fit_score", 0) >= 70 else "skip"

        with open(f"outputs/{company_name}.md", "w") as f:
            f.write(f"# {result.get('company')} — {result.get('position')}\n\n")
            f.write(f"**Fit Score:** {result.get('fit_score')}/100\n")
            f.write(f"**Decision:** {final_decision.upper()}\n\n")
            f.write(f"## Reasoning\n{result.get('reasoning', '')}\n\n")
            f.write(f"## Matches\n")
            for m in result.get("matches", []):
                f.write(f"- {m}\n")
            f.write(f"\n## Gaps\n")
            for g in result.get("gaps", []):
                f.write(f"- {g}\n")
            if final_decision == "apply":
                f.write(f"\n## CV Recommendations\n")
                f.write(f"**Add:** {', '.join(result.get('cv_recommendations', {}).get('add', []))}\n")
                f.write(f"**Remove:** {', '.join(result.get('cv_recommendations', {}).get('remove', []))}\n")
        
        if final_decision == "apply":
            save_to_tracker(result, url)
            cv_improvements = improve_cv(result, "data/cv.md", job_text)
            improvement_file = f"outputs/{company_name}_cv.md"
            with open(improvement_file, "w") as f:
                f.write(f"# CV Improvements — {result.get('company')} {result.get('position')}\n\n")
                f.write(cv_improvements)
            print(f"✓ {result.get('company')} — {result.get('fit_score')}/100 — APPLY → saved to tracker")
            print(f"  CV improvements: {improvement_file}")
        else:
            print(f"✗ {result.get('company')} — {result.get('fit_score')}/100 — SKIP")

        
    except Exception as e:
        import traceback
        traceback.print_exc()